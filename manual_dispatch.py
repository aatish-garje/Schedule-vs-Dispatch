import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

st.set_page_config(layout="wide")
st.title("Schedule vs Dispatch Report")

# --- Google Drive Links (still used for schedule & kit) ---
sales_register_url = "https://drive.google.com/uc?id=19pNz65xMPEBCrXQuEiFu6VVDXYhqeYjy"  # NOT used now (manual upload only)
schedule_url = "https://drive.google.com/uc?id=19FkajdpPaiQHqXqR5eH0WqximI5Sohs7"
kit_part_url = "https://drive.google.com/uc?id=18YkiGvirKsrrwg8IZq2H3Aje5HAw-Djp"
fg_url = "https://drive.google.com/uc?id=1JIEDeDMXOQKyxCUJTv2zHp6_hUecXBIt"  # kept just as reference; not auto-loaded

# --- File upload / selection controls ---

# 1) Sales Register: ALWAYS manual upload
dispatch_file = st.file_uploader(
    "Upload Sales Register (Dispatch) Excel file",
    type=["xlsx", "xls"],
    key="dispatch_file"
)

# 2) Schedule: option to use Google Drive or manual upload
st.sidebar.markdown("### Schedule File Source")
schedule_source = st.sidebar.radio(
    "Select Schedule Source",
    ["Use Google Drive file", "Upload schedule manually"],
    index=0
)

uploaded_schedule_file = None
if schedule_source == "Upload schedule manually":
    uploaded_schedule_file = st.sidebar.file_uploader(
        "Upload Schedule Excel (with POWER & MECH sheets)",
        type=["xlsx", "xls"],
        key="schedule_file"
    )

# 3) FG Stock: optional upload
fg_file = st.sidebar.file_uploader(
    "Upload FG Stock File (optional)",
    type=["xlsx", "xls"],
    key="fg_file"
)

# --- Block execution until mandatory files are provided ---

if dispatch_file is None:
    st.warning("Please upload the Sales Register (Dispatch) Excel file to continue.")
    st.stop()

# Determine schedule file object
if schedule_source == "Use Google Drive file":
    try:
        schedule_file = BytesIO(requests.get(schedule_url).content)
    except Exception as e:
        st.error(f"Error loading schedule from Google Drive: {e}")
        st.stop()
else:
    if uploaded_schedule_file is None:
        st.warning("Please upload the Schedule Excel file to continue.")
        st.stop()
    schedule_file = uploaded_schedule_file

# FG availability flag
fg_available = fg_file is not None
if fg_available:
    try:
        fg_df = pd.read_excel(fg_file)
    except Exception as e:
        st.error(f"Error reading uploaded FG file: {e}")
        st.stop()
else:
    fg_df = None

# --- Load data ---

# Sales register (dispatch) from manual upload
dispatch_df = pd.read_excel(dispatch_file)

# Schedule sheets from selected source
schedule_power = pd.read_excel(schedule_file, sheet_name="POWER", header=3)
schedule_mech = pd.read_excel(schedule_file, sheet_name="MECH", header=3)

# Kit file always from Google Drive
kit_file = BytesIO(requests.get(kit_part_url).content)

# --- Prepare kit lookups ---
kit_psg = pd.read_excel(kit_file, sheet_name="PSG", usecols="K:M")
kit_psg.columns = ["Part Number", "Desc", "Kit Part Number"]
kit_psg["Part Number"] = kit_psg["Part Number"].astype(str)
lookup_power_stg = dict(zip(kit_psg["Part Number"], kit_psg["Kit Part Number"]))

kit_psg_mech = pd.read_excel(kit_file, sheet_name="PSG", usecols="S:T")
kit_psg_mech.columns = ["Part Number", "Kit Part Number"]
kit_psg_mech["Part Number"] = kit_psg_mech["Part Number"].astype(str)
lookup_mech = dict(zip(kit_psg_mech["Part Number"], kit_psg_mech["Kit Part Number"]))

kit_vp = pd.read_excel(kit_file, sheet_name="VP", usecols="B:D")
kit_vp.columns = ["Part Number", "Desc", "Kit Part Number"]
kit_vp["Part Number"] = kit_vp["Part Number"].astype(str)
lookup_power_vp = dict(zip(kit_vp["Part Number"], kit_vp["Kit Part Number"]))

# --- Ensure Sold-to Party is string for safe comparisons ---
dispatch_df['Sold-to Party'] = dispatch_df['Sold-to Party'].astype(str)

billing_type_candidates = ['Billing Doc type', 'Billing Doc Type']
billing_type_col = next((c for c in billing_type_candidates if c in dispatch_df.columns), None)

if billing_type_col is not None:
    # 1. Keep only F2 and S1
    dispatch_df = dispatch_df[dispatch_df[billing_type_col].isin(['F2', 'S1'])].copy()

    required_cols = [billing_type_col, 'Cancel Doc', 'Billing Doc No.', 'Material']

    if all(col in dispatch_df.columns for col in required_cols):

        # Clean data (remove .0, spaces, ensure string)
        dispatch_df['Cancel Doc'] = (
            dispatch_df['Cancel Doc']
            .astype(str).str.strip()
            .str.replace(r'\.0$', '', regex=True)
        )
        dispatch_df['Billing Doc No.'] = (
            dispatch_df['Billing Doc No.']
            .astype(str).str.strip()
            .str.replace(r'\.0$', '', regex=True)
        )
        dispatch_df['Material'] = (
            dispatch_df['Material']
            .astype(str).str.strip()
            .str.replace(r'\.0$', '', regex=True)
        )

        # A. S1 rows (cancellation invoices)
        s1_rows = dispatch_df[dispatch_df[billing_type_col] == 'S1'].copy()

        # Optional: consider only S1 rows that actually have a Cancel Doc value
        s1_rows = s1_rows[
            (s1_rows['Cancel Doc'] != '') &
            (s1_rows['Cancel Doc'].str.lower() != 'nan')
        ]

        # B. Build cancellation keys from S1: (Cancel Doc, Material)
        cancellation_keys = set(zip(s1_rows['Cancel Doc'], s1_rows['Material']))

        # C. Row-wise filter logic
        def filter_invoices(row):
            # 1) Always remove S1 rows
            if row[billing_type_col] == 'S1':
                return False

            # 2) For F2 rows: drop if they match (Billing Doc No., Material) in cancellation keys
            key = (row['Billing Doc No.'], row['Material'])
            if key in cancellation_keys:
                return False

            # 3) Keep all other F2 rows
            return True

        dispatch_df = dispatch_df[dispatch_df.apply(filter_invoices, axis=1)].copy()

# --- Special Handling for Sold-to Party Q0001 (set Customer Group = 10) ---
dispatch_df.loc[dispatch_df['Sold-to Party'] == 'Q0001', 'Customer Group'] = 10

# --- Updated Customer Code Logic: Add "." only if Plant = 2000 AND Sold-to Party starts with A or F ---
def normalize_sold_to(row):
    sold = str(row.get('Sold-to Party', '')) if pd.notna(row.get('Sold-to Party', '')) else ''
    plant = str(row.get('Plant', '')).strip()

    # Only apply when Plant = 2000 AND Sold-to Party starts with A or F (case-insensitive)
    if plant.startswith("2000") and sold.upper().startswith(('A', 'F')):
        if not sold.endswith('.'):
            return sold + '.'
    return sold

dispatch_df['Sold-to Party'] = dispatch_df.apply(normalize_sold_to, axis=1)

# --- Filter out C* materials and replace zero Inv Qty with Kit Qty ---
dispatch_df = dispatch_df[~dispatch_df['Material'].astype(str).str.startswith('C')]
if 'Inv Qty' in dispatch_df.columns and 'Kit Qty' in dispatch_df.columns:
    dispatch_df.loc[dispatch_df['Inv Qty'] == 0, 'Inv Qty'] = dispatch_df['Kit Qty']
dispatch_df = dispatch_df[dispatch_df['Material'] != 8043975905]

# --- Keep only Customer Group 10 (after Q0001 override above) ---
dispatch_df = dispatch_df[dispatch_df['Customer Group'] == 10]

# --- Select sales orders starting with 10 and handle duplicates logic ---
keep_sales_order_10 = dispatch_df[dispatch_df['Sales Order No'].astype(str).str.startswith('10')]
remaining = dispatch_df[~dispatch_df['Sales Order No'].astype(str).str.startswith('10')]

# Defensive duplicate detection: support common column name variants
billing_col_candidates = ['Billing Doc No.', 'Billing Doc No', 'Billing Doc.', 'Billing Doc', 'BillingDocNo']
billing_col = next((c for c in billing_col_candidates if c in remaining.columns), None)

if billing_col:
    duplicates = remaining[billing_col].value_counts()[lambda x: x > 1].index.tolist()
    duplicates_df = remaining[remaining[billing_col].isin(duplicates)]
    unique_df = remaining[~remaining[billing_col].isin(duplicates)]
else:
    # if billing doc col not found, assume no duplicates
    duplicates_df = pd.DataFrame(columns=remaining.columns)
    unique_df = remaining.copy()

# Keep only rows with Item == 10 in duplicates (if Item exists)
if 'Item' in duplicates_df.columns:
    duplicates_df = duplicates_df[duplicates_df['Item'] == 10]

dispatch_df = pd.concat([keep_sales_order_10, unique_df, duplicates_df], ignore_index=True)

# --- Dispatch summary (Sold-to Party, Material) aggregated ---
dispatch_summary = dispatch_df.groupby(['Sold-to Party', 'Material'], as_index=False)['Inv Qty'].sum()
dispatch_summary.rename(columns={'Inv Qty': 'Dispatch Qty'}, inplace=True)

# --- Ensure schedule types are strings for merges ---
for df in [schedule_power, schedule_mech]:
    if 'Code' in df.columns:
        df['Code'] = df['Code'].astype(str)
    if 'Part Number' in df.columns:
        df['Part Number'] = df['Part Number'].astype(str)

# --- Merge dispatch summary into schedules ---
schedule_power = pd.merge(
    schedule_power,
    dispatch_summary,
    left_on=['Code', 'Part Number'],
    right_on=['Sold-to Party', 'Material'],
    how='left'
)
schedule_mech = pd.merge(
    schedule_mech,
    dispatch_summary,
    left_on=['Code', 'Part Number'],
    right_on=['Sold-to Party', 'Material'],
    how='left'
)

schedule_power['Dispatch Qty'] = schedule_power['Dispatch Qty'].fillna(0)
schedule_mech['Dispatch Qty'] = schedule_mech['Dispatch Qty'].fillna(0)

# --- Kit part number logic ---
def get_power_kit(row):
    desc = str(row.get('Description', '')).strip()
    part = str(row.get('Part Number', ''))
    if desc in ['STG GEAR KIT', 'STG GEAR KIT H-Pas']:
        return lookup_power_stg.get(part, '')
    if 'VANE PUMP KIT' in desc:
        return lookup_power_vp.get(part, '')
    return ''

schedule_power.insert(
    schedule_power.columns.get_loc('Part Number') + 1,
    'Kit Part Number',
    schedule_power.apply(get_power_kit, axis=1)
)
schedule_mech.insert(
    schedule_mech.columns.get_loc('Part Number') + 1,
    'Kit Part Number',
    schedule_mech.apply(
        lambda row: lookup_mech.get(str(row.get('Part Number', '')), '')
        if str(row.get('Part Number', '')).startswith(('7820975', '734097')) else '',
        axis=1
    )
)

# --- FG preparation (ONLY if FG file uploaded) ---
if fg_available:
    fg_df['Material'] = fg_df['Material'].astype(str)
    fg_df['Plant'] = fg_df['Plant'].astype(str)

    def fg_sum(df, part_col, plant_col):
        def get_plant_value(row, plant_key):
            if plant_key in row:
                return row[plant_key]
            for alt in ['BILLING PLANT', 'Billing Plant', 'Billing_Plant']:
                if alt in row:
                    return row[alt]
            return ''
        def row_sum(row):
            part = str(row.get(part_col, '')).strip()
            plant = str(get_plant_value(row, plant_col)).strip()
            if part == '' or plant == '':
                return 0
            return fg_df.loc[(fg_df['Material'] == part) & (fg_df['Plant'] == plant), 'Unrestricted'].sum()
        return df.apply(lambda r: row_sum(r), axis=1)

    schedule_power['FG'] = fg_sum(schedule_power, 'Part Number', 'BILLING PLANT') + fg_sum(schedule_power, 'Kit Part Number', 'BILLING PLANT')
    schedule_mech['FG'] = fg_sum(schedule_mech, 'Part Number', 'Billing Plant') + fg_sum(schedule_mech, 'Kit Part Number', 'Billing Plant')

# --- Marketing columns detection ---
marketing_columns_power = [col for col in schedule_power.columns if str(col).startswith('Marketing Requirement')]
marketing_columns_mech = [col for col in schedule_mech.columns if str(col).startswith('Marketing Requirement')]

# --- Balance & Excess Dispatch calculations ---
marketing_sum_power = schedule_power[marketing_columns_power].sum(axis=1) if len(marketing_columns_power) > 0 else pd.Series(0, index=schedule_power.index)
marketing_sum_mech = schedule_mech[marketing_columns_mech].sum(axis=1) if len(marketing_columns_mech) > 0 else pd.Series(0, index=schedule_mech.index)

schedule_power['Balance Dispatch'] = (marketing_sum_power - schedule_power['Dispatch Qty']).clip(lower=0)
schedule_mech['Balance Dispatch'] = (marketing_sum_mech - schedule_mech['Dispatch Qty']).clip(lower=0)

schedule_power['Excess Dispatch'] = (schedule_power['Dispatch Qty'] - marketing_sum_power).clip(lower=0)
schedule_mech['Excess Dispatch'] = (schedule_mech['Dispatch Qty'] - marketing_sum_mech).clip(lower=0)

# --- Dispatchable FG: allocate ONLY if FG is available ---
def allocate_dispatchable_fg(df, part_col='Part Number', fg_col='FG', balance_col='Balance Dispatch', out_col='Dispatchable FG'):
    grouped = df.groupby(part_col).groups
    for part, idxs in grouped.items():
        idxs_list = list(idxs)
        idxs_list.sort(
            key=lambda i: float(df.at[i, balance_col]) if pd.notna(df.at[i, balance_col]) else 0.0,
            reverse=True
        )
        # Use single FG value (max) per part
        fg_series = df.loc[idxs_list, fg_col].fillna(0).astype(float)
        total_fg = float(fg_series.max()) if not fg_series.empty else 0.0
        remaining = float(total_fg)
        if remaining <= 0:
            continue
        for i in idxs_list:
            bal = float(df.at[i, balance_col]) if pd.notna(df.at[i, balance_col]) else 0.0
            if bal <= 0 or remaining <= 0:
                continue
            alloc = bal if remaining >= bal else remaining
            if alloc > 0:
                if float(alloc).is_integer():
                    df.at[i, out_col] = int(alloc)
                else:
                    df.at[i, out_col] = alloc
                remaining -= alloc
    return df

if fg_available:
    schedule_power['Dispatchable FG'] = 0
    schedule_mech['Dispatchable FG'] = 0
    schedule_power = allocate_dispatchable_fg(
        schedule_power,
        part_col='Part Number',
        fg_col='FG',
        balance_col='Balance Dispatch',
        out_col='Dispatchable FG'
    )
    schedule_mech = allocate_dispatchable_fg(
        schedule_mech,
        part_col='Part Number',
        fg_col='FG',
        balance_col='Balance Dispatch',
        out_col='Dispatchable FG'
    )

# --- Final column selection & ordering ---
# Power
power_base_cols = ['Code', 'Customer', 'MODEL', 'BILLING PLANT', 'Part Number', 'Kit Part Number',
                   'Customer Part', 'Description', 'Initial Schedule', 'REV-1', 'REV-2']
power_cols = [c for c in power_base_cols if c in schedule_power.columns]
power_cols += marketing_columns_power
for c in ['Dispatch Qty', 'Balance Dispatch', 'Excess Dispatch']:
    if c in schedule_power.columns:
        power_cols.append(c)
if fg_available:
    # Add FG-related columns only if FG uploaded
    if 'FG' in schedule_power.columns:
        power_cols.insert(power_cols.index('Excess Dispatch'), 'FG')
    if 'Dispatchable FG' in schedule_power.columns:
        power_cols.insert(power_cols.index('Excess Dispatch'), 'Dispatchable FG')
# ZFI SCOPE if present
if 'ZFI SCOPE' in schedule_power.columns:
    power_cols.append('ZFI SCOPE')

schedule_power = schedule_power[power_cols]

# Mech
mech_base_cols = ['Code', 'Customer', 'Model', 'Billing Plant', 'Part Number', 'Kit Part Number',
                  'Customer Part', 'Description', 'Initial Schedule', 'REV-1', 'REV-2']
mech_cols = [c for c in mech_base_cols if c in schedule_mech.columns]
mech_cols += marketing_columns_mech
for c in ['Dispatch Qty', 'Balance Dispatch', 'Excess Dispatch']:
    if c in schedule_mech.columns:
        mech_cols.append(c)
if fg_available:
    if 'FG' in schedule_mech.columns:
        mech_cols.insert(mech_cols.index('Excess Dispatch'), 'FG')
    if 'Dispatchable FG' in schedule_mech.columns:
        mech_cols.insert(mech_cols.index('Excess Dispatch'), 'Dispatchable FG')

schedule_mech = schedule_mech[mech_cols]

# --- UI: Choose view ---
view_option = st.sidebar.radio("Select View", ["All", "Power Schedule", "Mech Schedule"])

def apply_filters(df, code, customer, billing_plant, model, part_number_search, sheet_type):
    if code:
        df = df[df['Code'].isin(code)]
    if customer:
        df = df[df['Customer'].isin(customer)]
    if billing_plant:
        plant_col = 'BILLING PLANT' if sheet_type == 'Power' else 'Billing Plant'
        if plant_col in df.columns:
            df = df[df[plant_col].isin(billing_plant)]
    if model:
        model_col = 'MODEL' if sheet_type == 'Power' else 'Model'
        if model_col in df.columns:
            df = df[df[model_col].isin(model)]
    if part_number_search:
        df = df[df['Part Number'].str.contains(part_number_search, case=False, na=False)]
    return df

def display_subtotals(df):
    if st.get_option("theme.base") == "dark":
        bg_color = "#262730"
        text_color = "#FFFFFF"
    else:
        bg_color = "#FFFFFF"
        text_color = "#000000"

    marketing_cols = [col for col in df.columns if str(col).startswith('Marketing Requirement')]
    subtotal_data = {col: df[col].sum() for col in marketing_cols}
    if 'Dispatch Qty' in df.columns:
        subtotal_data['Dispatch Qty'] = df['Dispatch Qty'].sum()
    if 'Balance Dispatch' in df.columns:
        subtotal_data['Balance Dispatch'] = df['Balance Dispatch'].sum()
    if 'Excess Dispatch' in df.columns:
        subtotal_data['Excess Dispatch'] = df['Excess Dispatch'].sum()
    # Only show FG totals if those columns exist (i.e., FG file uploaded)
    if 'FG' in df.columns:
        subtotal_data['FG'] = df['FG'].sum()
    if 'Dispatchable FG' in df.columns:
        subtotal_data['Dispatchable FG'] = pd.to_numeric(df['Dispatchable FG'], errors='coerce').fillna(0).sum()

    st.markdown(
        f"""
        <div style="background-color:{bg_color}; padding:10px; border-radius:5px;">
            {"".join([f"<p style='color:{text_color};'>{k}: {v:.0f}</p>" for k, v in subtotal_data.items()])}
        </div>
        """,
        unsafe_allow_html=True
    )

# --- Views and filters ---
if view_option == "Power Schedule":
    code = st.sidebar.multiselect('Code', schedule_power['Code'].unique())
    customer = st.sidebar.multiselect('Customer', schedule_power['Customer'].unique())
    billing_plant = st.sidebar.multiselect(
        'Billing Plant',
        schedule_power['BILLING PLANT'].unique() if 'BILLING PLANT' in schedule_power.columns else []
    )
    model = st.sidebar.multiselect(
        'Model',
        schedule_power['MODEL'].unique() if 'MODEL' in schedule_power.columns else []
    )
    part_number_search = st.sidebar.text_input('Part Number (Type & Press Enter)')
    filtered_power = apply_filters(schedule_power, code, customer, billing_plant, model, part_number_search, 'Power')

    display_subtotals(filtered_power)
    st.dataframe(filtered_power, use_container_width=True)
    power_to_download = filtered_power
    mech_to_download = pd.DataFrame()

elif view_option == "Mech Schedule":
    code = st.sidebar.multiselect('Code', schedule_mech['Code'].unique())
    customer = st.sidebar.multiselect('Customer', schedule_mech['Customer'].unique())
    billing_plant = st.sidebar.multiselect(
        'Billing Plant',
        schedule_mech['Billing Plant'].unique() if 'Billing Plant' in schedule_mech.columns else []
    )
    model = st.sidebar.multiselect(
        'Model',
        schedule_mech['Model'].unique() if 'Model' in schedule_mech.columns else []
    )
    part_number_search = st.sidebar.text_input('Part Number (Type & Press Enter)')
    filtered_mech = apply_filters(schedule_mech, code, customer, billing_plant, model, part_number_search, 'Mech')

    display_subtotals(filtered_mech)
    st.dataframe(filtered_mech, use_container_width=True)
    power_to_download = pd.DataFrame()
    mech_to_download = filtered_mech

else:
    power_to_download = schedule_power.copy()
    mech_to_download = schedule_mech.copy()
    st.write("### Power Schedule")
    st.dataframe(schedule_power, use_container_width=True)
    st.write("### Mech Schedule")
    st.dataframe(schedule_mech, use_container_width=True)

# --- Download logic (Excel with thin borders and auto column widths) ---
if not power_to_download.empty or not mech_to_download.empty:
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not power_to_download.empty:
            power_to_download.to_excel(writer, sheet_name='Power', index=False)
        if not mech_to_download.empty:
            mech_to_download.to_excel(writer, sheet_name='Mech', index=False)
        workbook = writer.book
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for sheet_name in writer.sheets:
            worksheet = workbook[sheet_name]
            for col in worksheet.columns:
                max_len = max(len(str(cell.value)) for cell in col if cell.value) + 2
                worksheet.column_dimensions[get_column_letter(col[0].column)].width = max_len
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column
            ):
                for cell in row:
                    cell.border = thin_border
    output.seek(0)
    st.download_button(
        "Download Excel",
        output,
        "Schedule_with_Dispatch.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
