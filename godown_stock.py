import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def write_sheet(ws, data):
    border_style = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    for customer, group in data.groupby('Name'):
        ws.append([])  # Blank row before customer header
        ws.append([customer])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=11)
        ws.append([])  # Blank row after customer header

        ws.append(['Code', 'Name', 'Inv No', 'Inv Date', 'Item Code', 'Item Desc',
                   'Qty', 'Amount', 'Days', 'GDN Receipt', 'ASN'])
        header_row = ws.max_row
        for col in range(1, 12):
            ws.cell(row=header_row, column=col).border = border_style

        for r in dataframe_to_rows(group, index=False, header=False):
            ws.append(r)
            for col in range(1, 12):
                ws.cell(row=ws.max_row, column=col).border = border_style

        ws.append([])  # Blank row after each group

    # Auto-fit columns
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            if not cell.value or cell.coordinate in ws.merged_cells:
                continue
            max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


# --- Streamlit App ---
st.set_page_config(layout="wide")
st.title('Pending Godown Stock Report Generator')

uploaded_file = st.file_uploader("Upload Raw Pending Godown Stock Excel", type=["xlsx"])
if uploaded_file:
    df = pd.read_excel(uploaded_file, sheet_name='Sheet1', header=1)
    df.columns = ['#', 'Code', 'Name', 'Inv No', 'Inv Date', 'Item Code', 'Item Desc',
                  'Qty', 'Amount', 'Days', 'GDN Receipt', 'ASN']
    df = df.drop('#', axis=1)
    df['Days'] = pd.to_numeric(df['Days'], errors='coerce')

    df_30_45 = df[(df['Days'] >= 30) & (df['Days'] <= 45)]
    df_46_60 = df[(df['Days'] >= 46) & (df['Days'] <= 60)]
    df_61_more = df[df['Days'] >= 61]

    st.header("30 - 45 Days Stock")
    st.dataframe(df_30_45)
    st.header("46 - 60 Days Stock")
    st.dataframe(df_46_60)
    st.header("61 & Above Days Stock")
    st.dataframe(df_61_more)

    output = BytesIO()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '30 - 45 Days'
    write_sheet(ws1, df_30_45)
    ws2 = wb.create_sheet('46 - 60 Days')
    write_sheet(ws2, df_46_60)
    ws3 = wb.create_sheet('61 & Above Days')
    write_sheet(ws3, df_61_more)
    wb.save(output)
    output.seek(0)

    st.download_button(
        label="📥 Download Pending Godown Stock Excel",
        data=output,
        file_name='Pending_Godown_Stock.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
