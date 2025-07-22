import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl import Workbook

# Material group codes as per your specification
power_codes = ['80339', '80379', '80349', '80439', '80469', '80489', '80499', '88439', 'M0339', 'M0439']
vane_pump_codes = ['76139', '76729', '76739', '76749', '76769', '76919']
mechanical_codes = ['73409', '78209']
bevel_gear_codes = ['78609']
drop_arm_codes = ['7325012', '7348012', '7363012', '7373012', '7379012']
oil_tank_codes = ['7632472', '7672472', '7632975501']

def matches_any(x, codes):
    return any(str(x).startswith(code) for code in codes)

def not_761395513799_and_513899(x):
    return str(x) not in ["7613955137/99", "7613955138/99"]

def drop_arm_no_slash(x):
    return '/' not in str(x)

def add_subtotal(df):
    df = df.copy()
    col7 = df.columns[:7]
    df = df.loc[:, col7]
    if 'Unrestricted' in col7:
        df['Unrestricted'] = pd.to_numeric(df['Unrestricted'], errors='coerce')
        subtotal = df['Unrestricted'].sum(min_count=1)
    else:
        subtotal = ''
    subtotal_row = {col: "" for col in col7}
    subtotal_row[col7[0]] = "Subtotal"
    subtotal_row['Unrestricted'] = subtotal
    df = pd.concat([df, pd.DataFrame([subtotal_row])], ignore_index=True)
    return df

def format_worksheet(ws):
    thin = Side(border_style="thin", color="000000")
    max_row = ws.max_row
    for i, row in enumerate(ws.iter_rows(), 1):
        is_header = (i == 1)
        is_subtotal = False
        for cell in row:
            if is_header:
                cell.font = Font(bold=True)
            if str(cell.value).strip().lower() == "subtotal":
                is_subtotal = True
        if is_subtotal:
            for c in row:
                c.font = Font(bold=True)
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            if is_header or is_subtotal:
                cell.alignment = Alignment(horizontal="center")
    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            length = len(str(cell.value)) if cell.value is not None else 0
            if length > max_length:
                max_length = length
        ws.column_dimensions[col_letter].width = max_length + 2

def to_excel(sheets):
    out = BytesIO()
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        format_worksheet(ws)
    wb.save(out)
    out.seek(0)
    return out

st.title("FG Stock Report")

uploaded_file = st.file_uploader("Upload your fg.XLSX file", type="xlsx")
if uploaded_file:
    df = pd.read_excel(uploaded_file, dtype=str)
    if 'Unrestricted' in df.columns:
        df['Unrestricted'] = pd.to_numeric(df['Unrestricted'], errors='coerce')
    col7 = df.columns[:7]

    sheets1 = {
        "Power": add_subtotal(df[df['Material'].apply(lambda x: matches_any(x, power_codes))][col7]),
        "Vane Pump": add_subtotal(
            df[
                df['Material'].apply(lambda x: matches_any(x, vane_pump_codes))
                & df['Material'].apply(not_761395513799_and_513899)
            ][col7]
        ),
        "Mechanical": add_subtotal(df[df['Material'].apply(lambda x: matches_any(x, mechanical_codes))][col7]),
        "Bevel Gear": add_subtotal(df[df['Material'].apply(lambda x: matches_any(x, bevel_gear_codes))][col7]),
        "Drop Arm": add_subtotal(
            df[
                df['Material'].apply(lambda x: matches_any(x, drop_arm_codes))
                & df['Material'].apply(drop_arm_no_slash)
            ][col7]
        ),
        "Oil Tank": add_subtotal(df[df['Material'].apply(lambda x: matches_any(x, oil_tank_codes))][col7]),
        "All FG": add_subtotal(df[df['Material'].apply(not_761395513799_and_513899)][col7])
    }
    out1 = to_excel(sheets1)
    st.download_button(
        label="Download ALL FG.xlsx",
        data=out1,
        file_name="ALL FG.xlsx",
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    df2000 = df[df['Plant'] == "2000"]
    sheets2 = {
        "Power": add_subtotal(df2000[df2000['Material'].apply(lambda x: matches_any(x, power_codes))][col7]),
        "Vane Pump": add_subtotal(
            df2000[
                df2000['Material'].apply(lambda x: matches_any(x, vane_pump_codes))
                & df2000['Material'].apply(not_761395513799_and_513899)
            ][col7]
        ),
    }
    out2 = to_excel(sheets2)
    st.download_button(
        label="Download 2000 Plant FG.xlsx",
        data=out2,
        file_name="2000 Plant FG.xlsx",
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    st.success("Download Your FG Files")
