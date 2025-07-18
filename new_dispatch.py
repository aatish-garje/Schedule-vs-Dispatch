import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import openpyxl

st.set_page_config(layout="wide")
st.title("Schedule vs Dispatch Report")

# Sidebar Main Filter
view_option = st.sidebar.radio("Select View", ["All", "Power Schedule", "Mech Schedule"])

# Upload Files
dispatch_file = st.sidebar.file_uploader("Upload Sales Register", type=["xlsx"])
schedule_file = st.sidebar.file_uploader("Upload Schedule File", type=["xlsx"])

def apply_filters(df, code, customer, billing_plant, model, part_number_search, sheet_type):
    if code:
        df = df[df['Code'].isin(code)]
    if customer:
        df = df[df['Customer'].isin(customer)]
    if billing_plant:
        plant_col = 'BILLING PLANT' if sheet_type == 'Power' else 'Billing Plant'
        df = df[df[plant_col].isin(billing_plant)]
    if model:
        model_col = 'MODEL' if sheet_type == 'Power' else 'Model'
        df = df[df[model_col].isin(model)]
    if part_number_search:
        df = df[df['Part Number'].str.contains(part_number_search, case=False, na=False)]
    return df


if dispatch_file and schedule_file:
    dispatch_df = pd.read_excel(dispatch_file)
    schedule_power = pd.read_excel(schedule_file, sheet_name="POWER", header=3)
    schedule_mech = pd.read_excel(schedule_file, sheet_name="MECH", header=3)

    # Dispatch Processing
    dispatch_df['Sold-to Party'] = dispatch_df.apply(
        lambda x: str(x['Sold-to Party']) + '.' if (x['Plant'] == 2000 and not str(x['Sold-to Party']).upper().startswith('V')) else x['Sold-to Party'],
        axis=1
    )
    dispatch_df = dispatch_df[~dispatch_df['Material'].astype(str).str.startswith('C')]
    dispatch_df.loc[dispatch_df['Inv Qty'] == 0, 'Inv Qty'] = dispatch_df['Kit Qty']
    dispatch_df = dispatch_df[dispatch_df['Material'] != 8043975905]
    dispatch_df = dispatch_df[dispatch_df['Customer Group'] == 10]

    keep_sales_order_10 = dispatch_df[dispatch_df['Sales Order No'].astype(str).str.startswith('10')]
    remaining = dispatch_df[~dispatch_df['Sales Order No'].astype(str).str.startswith('10')]
    billing_doc_counts = remaining['Billing Doc No.'].value_counts()
    duplicates = billing_doc_counts[billing_doc_counts > 1].index.tolist()

    duplicates_df = remaining[remaining['Billing Doc No.'].isin(duplicates)]
    unique_df = remaining[~remaining['Billing Doc No.'].isin(duplicates)]
    duplicates_df = duplicates_df[duplicates_df['Item'] == 10]
    dispatch_df = pd.concat([keep_sales_order_10, unique_df, duplicates_df], ignore_index=True)

    dispatch_summary = dispatch_df.groupby(['Sold-to Party', 'Material'], as_index=False)['Inv Qty'].sum()
    dispatch_summary.rename(columns={'Inv Qty': 'Dispatch Qty'}, inplace=True)

    schedule_power['Code'] = schedule_power['Code'].astype(str)
    schedule_power['Part Number'] = schedule_power['Part Number'].astype(str)
    schedule_mech['Code'] = schedule_mech['Code'].astype(str)
    schedule_mech['Part Number'] = schedule_mech['Part Number'].astype(str)

    # Merge Dispatch Qty
    schedule_power = pd.merge(
        schedule_power, dispatch_summary,
        left_on=['Code', 'Part Number'],
        right_on=['Sold-to Party', 'Material'],
        how='left'
    )
    schedule_mech = pd.merge(
        schedule_mech, dispatch_summary,
        left_on=['Code', 'Part Number'],
        right_on=['Sold-to Party', 'Material'],
        how='left'
    )

    schedule_power['Dispatch Qty'] = schedule_power['Dispatch Qty'].fillna(0)
    schedule_mech['Dispatch Qty'] = schedule_mech['Dispatch Qty'].fillna(0)

    # --- Columns Cleanup ---
    columns_to_keep_power = [
        'Code', 'Customer', 'MODEL', 'BILLING PLANT', 'Part Number',
        'Customer Part', 'Description', 'Initial Schedule', 'REV-1', 'REV-2'
    ]
    marketing_columns_power = [col for col in schedule_power.columns if str(col).startswith('Marketing Requirement')]
    final_columns_power = columns_to_keep_power + marketing_columns_power + ['Dispatch Qty']
    schedule_power = schedule_power[final_columns_power]

    columns_to_keep_mech = [
        'Code', 'Customer', 'Model', 'Billing Plant', 'Part Number',
        'Customer Part', 'Description', 'Initial Schedule', 'REV-1', 'REV-2'
    ]
    marketing_columns_mech = [col for col in schedule_mech.columns if str(col).startswith('Marketing Requirement')]
    final_columns_mech = columns_to_keep_mech + marketing_columns_mech + ['Dispatch Qty']
    schedule_mech = schedule_mech[final_columns_mech]

    # --- Balance & Excess Columns ---
    for df, marketing_columns in [(schedule_power, marketing_columns_power), (schedule_mech, marketing_columns_mech)]:
        df['Balance Dispatch'] = (df[marketing_columns].sum(axis=1) - df['Dispatch Qty']).clip(lower=0)
        df['Excess Dispatch'] = df.apply(
            lambda x: x['Dispatch Qty'] - x[marketing_columns].sum() if x['Dispatch Qty'] > x[marketing_columns].sum() else 0,
            axis=1
        )

    # --- Sidebar Filters ---
    if view_option == "Power Schedule":
        st.header("Power Schedule")
        code = st.sidebar.multiselect('Code', schedule_power['Code'].unique())
        customer = st.sidebar.multiselect('Customer', schedule_power['Customer'].unique())
        billing_plant = st.sidebar.multiselect('Billing Plant', schedule_power['BILLING PLANT'].unique())
        model = st.sidebar.multiselect('Model', schedule_power['MODEL'].unique())
        part_number_search = st.sidebar.text_input('Part Number (Type & Press Enter)')
        filtered_power = apply_filters(schedule_power, code, customer, billing_plant, model, part_number_search, sheet_type='Power')

        if filtered_power.empty:
            st.warning("No data or Wrong Filter Selection")
        else:
            st.dataframe(filtered_power, use_container_width=True)

        power_to_download = filtered_power
        mech_to_download = pd.DataFrame()

    elif view_option == "Mech Schedule":
        st.header("Mech Schedule")
        code = st.sidebar.multiselect('Code', schedule_mech['Code'].unique())
        customer = st.sidebar.multiselect('Customer', schedule_mech['Customer'].unique())
        billing_plant = st.sidebar.multiselect('Billing Plant', schedule_mech['Billing Plant'].unique())
        model = st.sidebar.multiselect('Model', schedule_mech['Model'].unique())
        part_number_search = st.sidebar.text_input('Part Number (Type & Press Enter)')
        filtered_mech = apply_filters(schedule_mech, code, customer, billing_plant, model, part_number_search, sheet_type='Mech')

        if filtered_mech.empty:
            st.warning("No data or Wrong Filter Selection")
        else:
            st.dataframe(filtered_mech, use_container_width=True)

        power_to_download = pd.DataFrame()
        mech_to_download = filtered_mech

    else:
        st.header("All Schedules (Power & Mech)")
        power_to_download = schedule_power.copy()
        mech_to_download = schedule_mech.copy()

        if power_to_download.empty:
            st.warning("Power Schedule: No data")
        else:
            st.subheader("Power Schedule")
            st.dataframe(power_to_download, use_container_width=True)

        if mech_to_download.empty:
            st.warning("Mech Schedule: No data")
        else:
            st.subheader("Mech Schedule")
            st.dataframe(mech_to_download, use_container_width=True)

    # --- Excel Export with AutoFit & Borders ---
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not power_to_download.empty:
            power_to_download.to_excel(writer, sheet_name='Power', index=False)
        if not mech_to_download.empty:
            mech_to_download.to_excel(writer, sheet_name='Mech', index=False)

        workbook = writer.book
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for sheet_name in writer.sheets:
            worksheet = workbook[sheet_name]
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                worksheet.column_dimensions[get_column_letter(column)].width = (max_length + 2)
            max_row, max_col = worksheet.max_row, worksheet.max_column
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.border = thin_border

    output.seek(0)

    # --- Download Button ---
    st.subheader("📥 Download Schedule vs Dispatch Excel")
    st.download_button(
        label="Download Excel File",
        data=output,
        file_name="Schedule_with_Dispatch.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("Please upload both Sales Register and Schedule files to continue.")
