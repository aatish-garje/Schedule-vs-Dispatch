import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import openpyxl

st.set_page_config(layout="wide")
st.title("Schedule vs Dispatch Report")

# Upload Files
dispatch_file = st.file_uploader("Upload Sales Register", type=["xlsx"])
schedule_file = st.file_uploader("Upload Schedule File", type=["xlsx"])

if dispatch_file and schedule_file:
    dispatch_df = pd.read_excel(dispatch_file)
    schedule_power = pd.read_excel(schedule_file, sheet_name="POWER", header=3)
    schedule_mech = pd.read_excel(schedule_file, sheet_name="MECH", header=3)

    # Dispatch Processing
    dispatch_df['Sold-to Party'] = dispatch_df.apply(
        lambda x: str(x['Sold-to Party']) + '.' if (x['Plant'] == 2000 and not str(x['Sold-to Party']).upper().startswith('V')) else x['Sold-to Party'],
        axis=1
    )
    dispatch_df = dispatch_df[~dispatch_df['Material'].astype(str).str.startswith('C')]
    dispatch_df.loc[dispatch_df['Inv Qty'] == 0, 'Inv Qty'] = dispatch_df['Kit Qty']
    dispatch_df = dispatch_df[dispatch_df['Material'] != 8043975905]
    dispatch_df = dispatch_df[dispatch_df['Customer Group'] == 10]

    keep_sales_order_10 = dispatch_df[dispatch_df['Sales Order No'].astype(str).str.startswith('10')]
    remaining = dispatch_df[~dispatch_df['Sales Order No'].astype(str).str.startswith('10')]
    billing_doc_counts = remaining['Billing Doc No.'].value_counts()
    duplicates = billing_doc_counts[billing_doc_counts > 1].index.tolist()

    duplicates_df = remaining[remaining['Billing Doc No.'].isin(duplicates)]
    unique_df = remaining[~remaining['Billing Doc No.'].isin(duplicates)]

    duplicates_df = duplicates_df[duplicates_df['Item'] == 10]

    dispatch_df = pd.concat([keep_sales_order_10, unique_df, duplicates_df], ignore_index=True)

    dispatch_summary = dispatch_df.groupby(['Sold-to Party', 'Material'], as_index=False)['Inv Qty'].sum()
    dispatch_summary.rename(columns={'Inv Qty': 'Dispatch Qty'}, inplace=True)

    schedule_power['Code'] = schedule_power['Code'].astype(str)
    schedule_power['Part Number'] = schedule_power['Part Number'].astype(str)

    schedule_mech['Code'] = schedule_mech['Code'].astype(str)
    schedule_mech['Part Number'] = schedule_mech['Part Number'].astype(str)

    # Merge Dispatch Qty
    schedule_power = pd.merge(
        schedule_power,
        dispatch_summary,
        left_on=['Code', 'Part Number'],
        right_on=['Sold-to Party', 'Material'],
        how='left'
    )
    schedule_mech = pd.merge(
        schedule_mech,
        dispatch_summary,
        left_on=['Code', 'Part Number'],
        right_on=['Sold-to Party', 'Material'],
        how='left'
    )

    schedule_power['Dispatch Qty'] = schedule_power['Dispatch Qty'].fillna(0)
    schedule_mech['Dispatch Qty'] = schedule_mech['Dispatch Qty'].fillna(0)

    # --- POWER Columns --- #
    columns_to_keep_power = [
        'Code', 'Customer', 'MODEL', 'BILLING PLANT', 'Part Number',
        'Customer Part', 'Description', 'Initial Schedule',
        'REV-1', 'REV-2'
    ]
    marketing_columns_power = [col for col in schedule_power.columns if str(col).startswith('Marketing Requirement')]
    final_columns_power = columns_to_keep_power + marketing_columns_power + ['Dispatch Qty']
    schedule_power = schedule_power[final_columns_power]

    # --- MECH Columns --- #
    columns_to_keep_mech = [
        'Code', 'Customer', 'Model', 'Billing Plant', 'Part Number',
        'Customer Part', 'Description', 'Initial Schedule',
        'REV-1', 'REV-2'
    ]
    marketing_columns_mech = [col for col in schedule_mech.columns if str(col).startswith('Marketing Requirement')]
    final_columns_mech = columns_to_keep_mech + marketing_columns_mech + ['Dispatch Qty']
    schedule_mech = schedule_mech[final_columns_mech]

    # --- ADD BALANCE & EXCESS DISPATCH --- #
    for df, marketing_columns in [(schedule_power, marketing_columns_power), (schedule_mech, marketing_columns_mech)]:
        df['Balance Dispatch'] = (df[marketing_columns].sum(axis=1) - df['Dispatch Qty']).clip(lower=0)
        df['Excess Dispatch'] = df.apply(
            lambda x: x['Dispatch Qty'] - x[marketing_columns].sum() if x['Dispatch Qty'] > x[marketing_columns].sum() else 0,
            axis=1
        )

    # --- Export to Excel with AutoFit Columns & Borders --- #
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        schedule_power.to_excel(writer, sheet_name='Power', index=False)
        schedule_mech.to_excel(writer, sheet_name='Mech', index=False)

        workbook = writer.book
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for sheet_name in ['Power', 'Mech']:
            worksheet = workbook[sheet_name]

            # AutoFit Columns
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

            # Borders Only for Data
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.border = thin_border

    output.seek(0)

    # --- Streamlit Download Button --- #
    st.subheader("📥 Download Schedule vs Dispatch Excel")
    st.download_button(
        label="Download Excel File",
        data=output,
        file_name="Schedule_with_Dispatch.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.success("File is ready for download ✅")

else:
    st.info("Please upload both Schedule and Dispatch files to continue.")
